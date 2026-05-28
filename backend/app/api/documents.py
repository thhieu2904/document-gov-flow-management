import io
from datetime import date, datetime, time, timedelta

import openpyxl
from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import get_current_user, require_manager
from app.core.email import email_assignment_created, email_document_deleted
from app.core.storage import get_storage_provider
from app.email_utils import send_and_log_task
from app.models import Document, DocumentAssignment, DocumentAttachment, DocumentComment, EmailLog, Notification, User, now_utc, VN_TZ
from app.schemas import AssignmentCreate, CommentCreate, DocumentCreate, DocumentUpdate
from app.services import (
    attachment_to_dict,
    document_assignments,
    ensure_can_view,
    ensure_manager_owner,
    notify,
    sync_document_status,
)

router = APIRouter(prefix="/documents", tags=["documents"])


def user_dict(user: User | None) -> dict | None:
    if not user:
        return None
    return {"id": user.id, "full_name": user.full_name, "email": user.email, "role": user.role, "department_id": user.department_id}


def assignment_dict(item: DocumentAssignment, assignee: User | None = None) -> dict:
    return {
        "id": item.id,
        "document_id": item.document_id,
        "assigned_by": item.assigned_by,
        "assignee_id": item.assignee_id,
        "assignee": user_dict(assignee or item.assignee),
        "instruction": item.instruction,
        "result_note": item.result_note,
        "priority": item.priority,
        "status": item.status,
        "due_at": item.due_at,
        "started_at": item.started_at,
        "submitted_at": item.submitted_at,
        "completed_at": item.completed_at,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def assignment_display_status(item: DocumentAssignment) -> str:
    if item.status == "completed":
        if item.due_at and item.completed_at and item.completed_at > item.due_at:
            return "completed_late"
        return "completed"
    if item.due_at and item.due_at < now_utc():
        return "overdue"
    if item.due_at and item.due_at <= now_utc() + timedelta(days=3):
        return "due_soon"
    return item.status


def assignment_matches_status(item: DocumentAssignment, selected: set[str]) -> bool:
    display_status = assignment_display_status(item)
    if "open" in selected and item.status != "completed":
        return True
    if "draft" in selected and item.status == "pending":
        return True
    return display_status in selected


def document_dict(doc: Document, assignments: list[DocumentAssignment] | None = None, my_assignment: DocumentAssignment | None = None) -> dict:
    assignments = assignments or []
    payload = {
        "id": doc.id,
        "title": doc.title,
        "code": doc.code,
        "summary": doc.summary,
        "priority": doc.priority,
        "status": doc.status,
        "display_status": derived_document_status(doc),
        "issued_at": doc.issued_at,
        "due_at": doc.due_at,
        "created_by": doc.created_by,
        "department_id": doc.department_id,
        "completed_at": doc.completed_at,
        "created_at": doc.created_at,
        "updated_at": doc.updated_at,
        "assignment_count": len(assignments),
        "completed_count": len([item for item in assignments if item.status == "completed"]),
    }
    if my_assignment:
        is_completed = my_assignment.status == "completed"
        payload.update(
            {
                "my_assignment_id": my_assignment.id,
                "my_assignment_status": my_assignment.status,
                "my_assignment_display_status": assignment_display_status(my_assignment),
                "my_assignment_due_at": my_assignment.due_at,
                "my_assignment_completed_at": my_assignment.completed_at,
                "my_assignment_progress": "1/1" if is_completed else "0/1",
            }
        )
    return payload


def comment_dict(item: DocumentComment) -> dict:
    return {
        "id": item.id,
        "document_id": item.document_id,
        "assignment_id": item.assignment_id,
        "user_id": item.user_id,
        "content": item.content,
        "created_at": item.created_at,
    }


def visible_document_ids_for_user(user: User):
    if user.role == "manager":
        return select(Document.id)
    return select(DocumentAssignment.document_id).where(DocumentAssignment.assignee_id == user.id)


def apply_document_search(query, search: str | None):
    if search:
        pattern = f"%{search.strip()}%"
        query = query.where(or_(Document.title.ilike(pattern), Document.code.ilike(pattern), Document.summary.ilike(pattern)))
    return query


def derived_document_status(doc: Document) -> str:
    if doc.status == "completed":
        if doc.due_at and doc.completed_at and doc.completed_at > doc.due_at:
            return "completed_late"
        return "completed"
    if doc.due_at and doc.due_at < now_utc():
        return "overdue"
    if doc.due_at and doc.due_at <= now_utc() + timedelta(days=3):
        return "due_soon"
    return doc.status


def period_bounds(period: str, anchor_date: date | None) -> tuple[datetime | None, datetime | None]:
    if period == "all":
        return None, None
    anchor = anchor_date or datetime.now(VN_TZ).date()
    if period == "week":
        start_date = anchor - timedelta(days=anchor.weekday())
        end_date = start_date + timedelta(days=7)
    elif period == "month":
        start_date = anchor.replace(day=1)
        end_date = anchor.replace(year=anchor.year + 1, month=1, day=1) if anchor.month == 12 else anchor.replace(month=anchor.month + 1, day=1)
    else:
        raise HTTPException(status_code=400, detail="Kỳ xem không hợp lệ")
    return datetime.combine(start_date, time.min).replace(tzinfo=VN_TZ), datetime.combine(end_date, time.min).replace(tzinfo=VN_TZ)


def in_period(doc: Document, start: datetime | None, end: datetime | None) -> bool:
    if not start or not end:
        return True
    values = [doc.due_at, doc.issued_at, doc.completed_at]
    return any(value is not None and start <= value < end for value in values)


def sort_documents(
    docs: list[Document],
    assignments_by_doc: dict[str, list[DocumentAssignment]],
    sort_by: str,
    sort_dir: str,
    my_assignments_by_doc: dict[str, DocumentAssignment] | None = None,
) -> list[Document]:
    direction = -1 if sort_dir == "desc" else 1

    def progress(doc: Document) -> float:
        if my_assignments_by_doc and doc.id in my_assignments_by_doc:
            return 1 if my_assignments_by_doc[doc.id].status == "completed" else 0
        assignments = assignments_by_doc.get(doc.id, [])
        return len([item for item in assignments if item.status == "completed"]) / len(assignments) if assignments else 0

    def value(doc: Document):
        my_assignment = my_assignments_by_doc.get(doc.id) if my_assignments_by_doc else None
        if sort_by == "status":
            if my_assignment:
                return assignment_display_status(my_assignment)
            return derived_document_status(doc)
        if sort_by == "progress":
            return progress(doc)
        if sort_by == "priority":
            if my_assignment:
                return {"urgent": 3, "high": 2, "normal": 1}.get(my_assignment.priority, 0)
            return {"urgent": 3, "high": 2, "normal": 1}.get(doc.priority, 0)
        if sort_by in {"issued_at", "due_at", "created_at"}:
            if my_assignment and sort_by == "due_at":
                return my_assignment.due_at or datetime.min.replace(tzinfo=now_utc().tzinfo)
            return getattr(doc, sort_by) or datetime.min.replace(tzinfo=now_utc().tzinfo)
        if sort_by in {"code", "title"}:
            return (getattr(doc, sort_by) or "").lower()
        return doc.updated_at or datetime.min.replace(tzinfo=now_utc().tzinfo)

    return sorted(docs, key=value, reverse=direction < 0)


def paginate(db: Session, query, page: int, size: int) -> dict:
    safe_page = max(page, 1)
    safe_size = min(max(size, 1), 100)
    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    docs = db.scalars(query.offset((safe_page - 1) * safe_size).limit(safe_size)).all()
    assignment_rows = db.scalars(select(DocumentAssignment).where(DocumentAssignment.document_id.in_([doc.id for doc in docs]))).all() if docs else []
    by_doc: dict[str, list[DocumentAssignment]] = {}
    for item in assignment_rows:
        by_doc.setdefault(item.document_id, []).append(item)
    return {"items": [document_dict(doc, by_doc.get(doc.id, [])) for doc in docs], "page": safe_page, "size": safe_size, "total": total}


@router.get("/export")
def export_documents(
    scope: str = "assigned_by_me",
    status: list[str] = Query(default=[]),
    priority: list[str] = Query(default=[]),
    search: str | None = None,
    period: str = "all",
    anchor_date: date | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sort_by: str = "due_at",
    sort_dir: str = "asc",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Document)
    if current_user.role == "manager":
        if scope == "my_tasks":
            query = query.where(Document.id.in_(select(DocumentAssignment.document_id).where(DocumentAssignment.assignee_id == current_user.id)))
    else:
        query = query.where(Document.id.in_(visible_document_ids_for_user(current_user)))
    query = apply_document_search(query, search)
    all_docs = db.scalars(query).all()
    assignment_rows = db.scalars(select(DocumentAssignment).where(DocumentAssignment.document_id.in_([doc.id for doc in all_docs]))).all() if all_docs else []
    by_doc: dict[str, list[DocumentAssignment]] = {}
    for item in assignment_rows:
        by_doc.setdefault(item.document_id, []).append(item)
    my_by_doc = {item.document_id: item for item in assignment_rows if item.assignee_id == current_user.id} if scope == "my_tasks" else {}
    
    if period == "custom":
        start = datetime.combine(start_date, time.min).replace(tzinfo=VN_TZ) if start_date else None
        end = datetime.combine(end_date, time.max).replace(tzinfo=VN_TZ) if end_date else None
    else:
        start, end = period_bounds(period, anchor_date)
        
    filtered = [doc for doc in all_docs if in_period(doc, start, end)]
    if status:
        selected = set(status)
        if scope == "my_tasks":
            filtered = [doc for doc in filtered if doc.id in my_by_doc and assignment_matches_status(my_by_doc[doc.id], selected)]
        else:
            filtered = [doc for doc in filtered if ("open" in selected and doc.status != "completed") or derived_document_status(doc) in selected]
    if priority:
        selected_priority = set(priority)
        if scope == "my_tasks":
            filtered = [doc for doc in filtered if doc.id in my_by_doc and my_by_doc[doc.id].priority in selected_priority]
        else:
            filtered = [doc for doc in filtered if doc.priority in selected_priority]
    sorted_docs = sort_documents(filtered, by_doc, sort_by, sort_dir, my_by_doc if scope == "my_tasks" else None)

    from openpyxl.styles import Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách văn bản"

    # Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "BÁO CÁO TÌNH HÌNH XỬ LÝ VĂN BẢN"
    title_cell.font = Font(name='Times New Roman', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Subtitle
    if period == "custom" and start and end:
        period_text = f"Từ ngày {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}"
    elif period == "week" and start and end:
        period_text = f"Tuần từ {start.strftime('%d/%m/%Y')} đến {(end - timedelta(days=1)).strftime('%d/%m/%Y')}"
    elif period == "month" and start:
        period_text = f"Tháng {start.strftime('%m/%Y')}"
    else:
        period_text = "Tất cả thời gian"

    ws['A2'] = f"Kỳ báo cáo: {period_text}"
    ws['A2'].font = Font(name='Times New Roman', size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='left')
    
    # Calculate stats
    total_docs = len(sorted_docs)
    if scope == "my_tasks":
        completed_docs = sum(1 for d in sorted_docs if my_by_doc.get(d.id) and my_by_doc[d.id].status == "completed")
        in_progress_docs = sum(1 for d in sorted_docs if my_by_doc.get(d.id) and my_by_doc[d.id].status in ["pending", "in_progress"])
        overdue_docs = sum(1 for d in sorted_docs if my_by_doc.get(d.id) and assignment_display_status(my_by_doc[d.id]) in ["overdue", "completed_late"])
    else:
        completed_docs = sum(1 for d in sorted_docs if d.status == "completed")
        in_progress_docs = sum(1 for d in sorted_docs if d.status == "in_progress")
        overdue_docs = sum(1 for d in sorted_docs if derived_document_status(d) in ["overdue", "completed_late"])

    # Stats
    ws['A4'] = f"Tổng số văn bản: {total_docs}   |   Hoàn tất: {completed_docs}   |   Đang thực hiện: {in_progress_docs}   |   Quá hạn/Trễ: {overdue_docs}"
    ws['A4'].font = Font(name='Times New Roman', size=12, bold=True)

    headers = [
        "Số hiệu", "Trích yếu", "Ngày tạo", "Ngày ban hành", "Hạn hoàn thành",
        "Tiến độ", "Trạng thái", "Độ ưu tiên", "Phân công"
    ]
    # We will append the headers manually at row 6
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="214B74", end_color="214B74", fill_type="solid")
        cell.font = Font(name='Times New Roman', color="FFFFFF", size=12, bold=True)
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['I'].width = 70

    def fmt_dt(dt):
        if not dt:
            return ""
        local_dt = dt.astimezone() if dt.tzinfo else dt
        return local_dt.strftime("%d/%m/%Y %H:%M")

    status_map = {
        "draft": "Chưa giao", "in_progress": "Đang thực hiện", "completed": "Hoàn tất",
        "completed_late": "Hoàn tất trễ hạn", "due_soon": "Sắp đến hạn", "overdue": "Quá hạn",
    }
    priority_map = {"normal": "Thường", "high": "Khẩn", "urgent": "Hỏa tốc"}
    assign_status_map = {"pending": "Chờ xử lý", "in_progress": "Đang làm", "completed": "Xong", "overdue": "Trễ"}

    status_color_map = {
        "Chưa giao": "64748B",
        "Đang thực hiện": "D97706",
        "Hoàn tất": "059669",
        "Hoàn tất trễ hạn": "059669",
        "Sắp đến hạn": "DC2626",
        "Quá hạn": "DC2626",
    }
    priority_color_map = {
        "Khẩn": "D97706",
        "Hỏa tốc": "DC2626",
    }

    for idx, doc in enumerate(sorted_docs, start=7):
        assignments = by_doc.get(doc.id, [])
        my_assignment = my_by_doc.get(doc.id)
        completed_count = len([a for a in assignments if a.status == "completed"])
        progress = "1/1" if my_assignment and my_assignment.status == "completed" else "0/1" if my_assignment else f"{completed_count}/{len(assignments)}"
        display_status_key = assignment_display_status(my_assignment) if my_assignment else derived_document_status(doc)
        display_status = status_map.get(display_status_key, assign_status_map.get(display_status_key, display_status_key))
        doc_priority = priority_map.get(my_assignment.priority if my_assignment else doc.priority, my_assignment.priority if my_assignment else doc.priority)

        assignees_text = "\n".join([f"- {a.assignee.full_name if a.assignee else '?'} ({assign_status_map.get(a.status, a.status)})" for a in assignments])

        row_data = [
            doc.code or "-",
            doc.title,
            fmt_dt(doc.created_at),
            fmt_dt(doc.issued_at),
            fmt_dt(doc.due_at),
            progress,
            display_status,
            doc_priority,
            assignees_text
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_num)
            cell.value = value
            cell.font = Font(name='Times New Roman', size=12)
            if col_num in [2, 9]: # Trích yếu and Phân công
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top')

        if display_status in status_color_map:
            ws.cell(row=idx, column=7).font = Font(name='Times New Roman', size=12, color=status_color_map[display_status], bold=True)
        if doc_priority in priority_color_map:
            ws.cell(row=idx, column=8).font = Font(name='Times New Roman', size=12, color=priority_color_map[doc_priority], bold=True)

    ws.auto_filter.ref = f"A6:I{6 + len(sorted_docs)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Bao_cao_van_ban_{datetime.now(VN_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("")
def list_documents(
    scope: str = "assigned_by_me",
    status: list[str] = Query(default=[]),
    priority: list[str] = Query(default=[]),
    search: str | None = None,
    period: str = "all",
    anchor_date: date | None = None,
    sort_by: str = "due_at",
    sort_dir: str = "asc",
    page: int = 1,
    size: int = 20,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Document)
    if current_user.role == "manager":
        if scope == "my_tasks":
            query = query.where(Document.id.in_(select(DocumentAssignment.document_id).where(DocumentAssignment.assignee_id == current_user.id)))
    else:
        query = query.where(Document.id.in_(visible_document_ids_for_user(current_user)))
    query = apply_document_search(query, search)
    all_docs = db.scalars(query).all()
    assignment_rows = db.scalars(select(DocumentAssignment).where(DocumentAssignment.document_id.in_([doc.id for doc in all_docs]))).all() if all_docs else []
    by_doc: dict[str, list[DocumentAssignment]] = {}
    for item in assignment_rows:
        by_doc.setdefault(item.document_id, []).append(item)
    my_by_doc = {item.document_id: item for item in assignment_rows if item.assignee_id == current_user.id} if scope == "my_tasks" else {}
    start, end = period_bounds(period, anchor_date)
    filtered = [doc for doc in all_docs if in_period(doc, start, end)]
    if status:
        selected = set(status)
        if scope == "my_tasks":
            filtered = [doc for doc in filtered if doc.id in my_by_doc and assignment_matches_status(my_by_doc[doc.id], selected)]
        else:
            filtered = [doc for doc in filtered if ("open" in selected and doc.status != "completed") or derived_document_status(doc) in selected]
    if priority:
        selected_priority = set(priority)
        if scope == "my_tasks":
            filtered = [doc for doc in filtered if doc.id in my_by_doc and my_by_doc[doc.id].priority in selected_priority]
        else:
            filtered = [doc for doc in filtered if doc.priority in selected_priority]
    sorted_docs = sort_documents(filtered, by_doc, sort_by, sort_dir, my_by_doc if scope == "my_tasks" else None)
    safe_page = max(page, 1)
    safe_size = min(max(size, 1), 100)
    start_index = (safe_page - 1) * safe_size
    page_docs = sorted_docs[start_index:start_index + safe_size]
    return {"items": [document_dict(doc, by_doc.get(doc.id, []), my_by_doc.get(doc.id)) for doc in page_docs], "page": safe_page, "size": safe_size, "total": len(sorted_docs)}


@router.post("")
def create_document(payload: DocumentCreate, db: Session = Depends(get_db), current_user: User = Depends(require_manager)):
    doc = Document(**payload.model_dump(), status="draft", created_by=current_user.id)
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return document_dict(doc)


@router.get("/{document_id}")
def get_document(document_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    assignments = document_assignments(db, doc.id)
    ensure_can_view(current_user, doc, assignments)
    comments = db.scalars(select(DocumentComment).where(DocumentComment.document_id == doc.id).order_by(DocumentComment.created_at)).all()
    attachments = db.scalars(select(DocumentAttachment).where(DocumentAttachment.document_id == doc.id).order_by(DocumentAttachment.created_at.desc())).all()
    uploader_ids = {item.uploaded_by for item in attachments}
    uploaders = {item.id: item for item in db.scalars(select(User).where(User.id.in_(uploader_ids))).all()} if uploader_ids else {}
    return {
        **document_dict(doc, assignments),
        "assignments": [assignment_dict(item) for item in assignments],
        "comments": [comment_dict(item) for item in comments],
        "attachments": [attachment_to_dict(item, uploaders.get(item.uploaded_by)) for item in attachments],
        "my_permissions": {
            "can_update": current_user.role == "manager",
            "can_assign": current_user.role == "manager",
            "can_delete": current_user.role == "manager",
        },
    }


@router.patch("/{document_id}")
def update_document(document_id: str, payload: DocumentUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_manager)):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    ensure_manager_owner(current_user, doc)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(doc, key, value)
    sync_document_status(db, doc)
    db.commit()
    db.refresh(doc)
    return document_dict(doc, document_assignments(db, doc.id))


@router.post("/{document_id}/assign")
def assign_document(
    document_id: str,
    payload: AssignmentCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager),
):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    ensure_manager_owner(current_user, doc)
    staff = db.scalars(select(User).where(User.id.in_(payload.assignee_ids), User.role == "staff", User.is_active.is_(True))).all()
    if len(staff) != len(set(payload.assignee_ids)):
        raise HTTPException(status_code=400, detail="Danh sách nhân viên không hợp lệ")
    existing = {
        item.assignee_id
        for item in db.scalars(select(DocumentAssignment).where(DocumentAssignment.document_id == doc.id)).all()
    }
    created = []
    for assignee in staff:
        if assignee.id in existing:
            continue
        assignment = DocumentAssignment(
            document_id=doc.id,
            assigned_by=current_user.id,
            assignee_id=assignee.id,
            instruction=payload.instruction,
            due_at=payload.due_at or doc.due_at,
            priority=payload.priority,
            status="pending",
        )
        db.add(assignment)
        db.flush()
        notify(db, assignee.id, doc.id, "Bạn được giao văn bản", doc.title, assignment.id)
        created.append(assignment)
    sync_document_status(db, doc)
    db.commit()
    if settings.email_enabled:
        for assignment in created:
            assignee = db.get(User, assignment.assignee_id)
            if not assignee:
                continue
            subject, html = email_assignment_created(
                doc.title,
                doc.code,
                payload.instruction,
                str(assignment.due_at or ""),
                settings.frontend_url,
            )
            background_tasks.add_task(
                send_and_log_task,
                log_key=f"assignment_created:{assignment.id}:{assignee.id}",
                event_type="assignment_created",
                recipient_email=assignee.email,
                subject=subject,
                html=html,
                document_id=doc.id,
                assignment_id=assignment.id,
                recipient_user_id=assignee.id,
            )
    return [assignment_dict(item) for item in created]


@router.delete("/{document_id}", status_code=204)
def delete_document(document_id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: User = Depends(require_manager)):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    ensure_manager_owner(current_user, doc)
    # Collect assigned staff for email notification before deleting
    assignments = document_assignments(db, doc.id)
    staff_to_notify = []
    if settings.email_enabled and assignments:
        assignee_ids = {a.assignee_id for a in assignments}
        staff_to_notify = list(db.scalars(select(User).where(User.id.in_(assignee_ids), User.is_active.is_(True))).all())
    doc_title = doc.title
    doc_code = doc.code
    assignment_ids = [item.id for item in assignments]
    attachments = db.scalars(select(DocumentAttachment).where(DocumentAttachment.document_id == doc.id)).all()
    for attachment in attachments:
        try:
            get_storage_provider(attachment.storage_provider).delete(attachment.storage_key)
        except Exception:
            pass
    email_log_filter = EmailLog.document_id == doc.id
    if assignment_ids:
        email_log_filter = or_(email_log_filter, EmailLog.assignment_id.in_(assignment_ids))
    db.query(EmailLog).filter(email_log_filter).update({EmailLog.document_id: None, EmailLog.assignment_id: None}, synchronize_session=False)
    db.query(Notification).filter(Notification.document_id == doc.id).delete(synchronize_session=False)
    db.query(DocumentComment).filter(DocumentComment.document_id == doc.id).delete(synchronize_session=False)
    db.query(DocumentAttachment).filter(DocumentAttachment.document_id == doc.id).delete(synchronize_session=False)
    db.query(DocumentAssignment).filter(DocumentAssignment.document_id == doc.id).delete(synchronize_session=False)
    db.delete(doc)
    db.commit()
    # Send deletion notification emails
    for staff in staff_to_notify:
        subject, html = email_document_deleted(doc_title, doc_code, staff.full_name)
        background_tasks.add_task(
            send_and_log_task,
            log_key=f"document_deleted:{doc.id}:{staff.id}",
            event_type="document_deleted",
            recipient_email=staff.email,
            subject=subject,
            html=html,
            recipient_user_id=staff.id,
        )
    return None


@router.post("/{document_id}/comments")
def add_comment(document_id: str, payload: CommentCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    assignments = document_assignments(db, doc.id)
    ensure_can_view(current_user, doc, assignments)
    comment = DocumentComment(document_id=doc.id, assignment_id=payload.assignment_id, user_id=current_user.id, content=payload.content)
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return comment_dict(comment)


@router.post("/{document_id}/attachments")
async def upload_attachment(
    document_id: str,
    file: UploadFile = File(...),
    assignment_id: str | None = Form(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.get(Document, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    assignments = document_assignments(db, doc.id)
    ensure_can_view(current_user, doc, assignments)
    # Manager can only upload original files when document has no assignments yet
    if current_user.role == "manager" and not assignment_id and assignments:
        raise HTTPException(status_code=400, detail="Văn bản đã giao việc, không thể thêm file gốc")
    if current_user.role == "staff" and not assignment_id:
        raise HTTPException(status_code=400, detail="Nhân viên chỉ được upload file kết quả cho việc được giao")
    if assignment_id:
        assignment = next((item for item in assignments if item.id == assignment_id), None)
        if not assignment:
            raise HTTPException(status_code=400, detail="Assignment không thuộc văn bản này")
        if current_user.role == "staff" and assignment.assignee_id != current_user.id:
            raise HTTPException(status_code=403, detail="Bạn không được upload file cho việc của người khác")
    storage_key, size = await get_storage_provider().save(file, f"documents/{doc.id}")
    attachment = DocumentAttachment(
        document_id=doc.id,
        assignment_id=assignment_id,
        storage_provider=settings.storage_provider,
        storage_key=storage_key,
        original_name=file.filename or "file",
        mime_type=file.content_type or "application/octet-stream",
        size=size,
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    return attachment_to_dict(attachment, current_user)
